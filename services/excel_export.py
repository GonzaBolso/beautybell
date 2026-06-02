from datetime import date
import os
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)


# Colores BeautyBel
ROSA        = "E91E8C"
ROSA_SUAVE  = "FCE4F3"
VERDE       = "D4EDDA"
ROJO        = "F8D7DA"
GRIS_CLARO  = "F7F4F6"
BLANCO      = "FFFFFF"
TEXTO_OSC   = "1A1A2E"


def _borde_fino():
    lado = Side(style="thin", color="E0D9E8")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _celda(ws, fila, col, valor, negrita=False, color_fondo=None,
           color_texto=None, alineacion="left", formato=None):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font = Font(
        name="Calibri", size=11,
        bold=negrita,
        color=color_texto or TEXTO_OSC
    )
    if color_fondo:
        c.fill = PatternFill("solid", fgColor=color_fondo)
    c.alignment = Alignment(horizontal=alineacion, vertical="center", wrap_text=False)
    c.border = _borde_fino()
    if formato:
        c.number_format = formato
    return c


def exportar_movimientos(movimientos: list, fecha_desde: str,
                         fecha_hasta: str, carpeta: str = None) -> str:
    """
    Genera un Excel con los movimientos de caja del periodo.
    Retorna la ruta del archivo generado.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Anchos de columna
    anchos = [14, 14, 20, 42, 18, 16]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = ancho

    fila = 1

    # ---- Título ----
    ws.merge_cells(f"A{fila}:F{fila}")
    c = ws.cell(row=fila, column=1, value="BeautyBel — Movimientos de Caja")
    c.font = Font(name="Calibri", size=16, bold=True, color=ROSA)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=ROSA_SUAVE)
    ws.row_dimensions[fila].height = 32
    fila += 1

    # ---- Periodo ----
    ws.merge_cells(f"A{fila}:F{fila}")
    c = ws.cell(row=fila, column=1,
                value=f"Periodo: {fecha_desde}  →  {fecha_hasta}")
    c.font = Font(name="Calibri", size=11, color=TEXTO_OSC)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", fgColor=GRIS_CLARO)
    ws.row_dimensions[fila].height = 20
    fila += 1

    fila += 1  # espacio

    # ---- Encabezado tabla ----
    encabezados = ["Fecha", "Tipo", "Categoria", "Descripcion", "Metodo de pago", "Monto"]
    for col, texto in enumerate(encabezados, 1):
        c = _celda(ws, fila, col, texto, negrita=True,
                   color_fondo=ROSA, color_texto=BLANCO, alineacion="center")
    ws.row_dimensions[fila].height = 22
    fila += 1

    # ---- Filas de datos ----
    total_ingresos = 0.0
    total_egresos  = 0.0

    for m in movimientos:
        es_ingreso = m["tipo"] == "ingreso"
        fondo_fila = VERDE if es_ingreso else ROJO
        monto_val  = m["monto"] if es_ingreso else -m["monto"]

        if es_ingreso:
            total_ingresos += m["monto"]
        else:
            total_egresos += m["monto"]

        desc_raw = m["descripcion"] or ""
        # Reformatear "Cobro: X - Servicio - Empleada" -> "Cobro: X - Servicio (Empleada)"
        if desc_raw.startswith("Cobro:") and desc_raw.count(" - ") >= 2:
            partes = desc_raw.split(" - ")
            empleada = partes[-1]
            desc_excel = " - ".join(partes[:-1]) + " (" + empleada + ")"
        else:
            desc_excel = desc_raw

        _celda(ws, fila, 1, m["fecha"] or "", color_fondo=fondo_fila)
        _celda(ws, fila, 2, m["tipo"].capitalize(), color_fondo=fondo_fila,
               alineacion="center")
        _celda(ws, fila, 3, m["categoria"] or "", color_fondo=fondo_fila)
        _celda(ws, fila, 4, desc_excel, color_fondo=fondo_fila)
        _celda(ws, fila, 5, m["metodo_pago_nombre"] or "—", color_fondo=fondo_fila,
               alineacion="center")
        _celda(ws, fila, 6, monto_val, color_fondo=fondo_fila,
               alineacion="right", formato='"$"#,##0.00')
        ws.row_dimensions[fila].height = 20
        fila += 1

    fila += 1  # espacio

    # ---- Resumen ----
    saldo = total_ingresos - total_egresos
    resumen = [
        ("Total ingresos", total_ingresos,  VERDE),
        ("Total egresos",  total_egresos,   ROJO),
        ("Saldo",          saldo, VERDE if saldo >= 0 else ROJO),
    ]
    for etiqueta, valor, fondo in resumen:
        ws.merge_cells(f"A{fila}:E{fila}")
        _celda(ws, fila, 1, etiqueta, negrita=True,
               color_fondo=fondo, alineacion="right")
        # Limpiar merge y poner valor en col F
        for col in range(2, 6):
            c = ws.cell(row=fila, column=col)
            c.fill = PatternFill("solid", fgColor=fondo)
        _celda(ws, fila, 6, valor, negrita=True,
               color_fondo=fondo, alineacion="right",
               formato='"$"#,##0.00')
        ws.row_dimensions[fila].height = 20
        fila += 1

    # ---- Pie ----
    fila += 1
    ws.merge_cells(f"A{fila}:F{fila}")
    c = ws.cell(row=fila, column=1,
                value=f"Generado el {date.today().strftime('%d/%m/%Y')}")
    c.font = Font(name="Calibri", size=9, color="888888")
    c.alignment = Alignment(horizontal="right")

    # ---- Guardar ----
    if not carpeta:
        carpeta = os.path.expanduser("~\\Documents") if os.name == "nt" \
                  else os.path.expanduser("~/Documents")
    os.makedirs(carpeta, exist_ok=True)

    nombre = f"BeautyBel_Caja_{fecha_desde}_al_{fecha_hasta}.xlsx"
    ruta   = os.path.normpath(os.path.join(carpeta, nombre))
    wb.save(ruta)
    return ruta